# -*- coding: utf-8 -*-
"""
Created on Wed Jul 29 07:57:54 2020

@author: papalk
"""
# Code logbook
# 10.02.2021 - Handover from Alkistis
# changed filename to filename[0] due to variable type error
# Changing from df.ix[...,0] to df.loc[..., df.columns[0]] due to a newer pandas version


import sys
import datetime as dt
import matplotlib.pyplot as plt
import matplotlib.pylab as pylab
from matplotlib.dates import DateFormatter
from matplotlib.ticker import StrMethodFormatter
import pandas as pd
import numpy as np
import xarray as xr
from mpl_toolkits.mplot3d import Axes3D
import glob 

# import user modules
usermodPath = r'../../userModules'
sys.path.append(usermodPath)
import pythonAssist
from pythonAssist import *

plt.style.use('seaborn-whitegrid')
SMALL_SIZE = 17
MEDIUM_SIZE = 22
BIGGER_SIZE = 22
AppendLog=1

plt.rc('font', size=SMALL_SIZE,weight = 'bold')          # controls default text sizes
plt.rc('axes', titlesize=SMALL_SIZE)     # fontsize of the axes title
plt.rc('axes', labelsize=BIGGER_SIZE)    # fontsize of the x and y labels
plt.rc('xtick', labelsize=MEDIUM_SIZE)    # fontsize of the tick labels
plt.rc('ytick', labelsize=MEDIUM_SIZE)    # fontsize of the tick labels
plt.rc('legend', fontsize=MEDIUM_SIZE)    # legend fontsize
plt.rc('figure', titlesize=BIGGER_SIZE)  # fontsize of the figure title
plt.rc('figure', figsize =  (8, 8))

#%% Import data
dt_start ='2021-08-02 00:00:00' # Select start date in the form yyyy-mm-dd_HH-MM-SS

dt_end = dt.datetime.strptime(dt_start, '%Y-%m-%d %H:%M:%S')  + dt.timedelta(days=7)# Select end date in the form yyyy-mm-dd_HH-MM-SS
dt_end = dt_end.strftime('%Y-%m-%d %H:%M:%S')

# Import csv
path = r'Z:\Projekte\109797-TestfeldBHV\30_Technical_execution_Confidential\TP3\AP2_Aufbau_Infrastruktur\Infrastruktur_Windmessung\02_Equipment\02_iSpin\Data\10min_Data'
filename = glob.glob(path+'\Bremerhaven WTG01*.csv')
df=pd.read_csv(filename[0], sep = ';',decimal = ',',header=0)

df['TimeStamp'] = [dt.datetime.strptime(date, '%Y-%m-%d %H:%M:%S') for date in df['TimeStamp']]

## Filtering conditions
cond0 = df['WS_free_avg'].notnull() # non-zero values
cond1 = (df["TimeStamp"]>=dt_start)&(df["TimeStamp"]<dt_end) # time interval filtering
cond2 = (df["SaDataValid"]==True) # 95% availability of 10 Hz data, wind vector +-90° in front of turbine within 10 min 
cond3 = (df["DataOK"]==True)  # data.TotalCountNoRotation=0, 95% availability, min & max rotor rpm, avg rotor rpm, free wind speed > 3.5 m/s, sample ID!= 0
cond4 = (df['WS_free_avg'] > 0) & (df['WS_free_avg'] < 50) # wind speed physical limits

## extra parameters for logbook
# no. of pts during the last week 
Npts = df.loc[(cond0 & cond1),:].shape[0]
Nvalid = df.loc[(cond0 & cond1 & cond2),:].shape[0]
Nws_valid = df.loc[(cond0 & cond1 & cond2),df.columns.values[1]].shape[0]
Nwt_valid = df.loc[(cond0 & cond1 & cond2 & cond3),:].shape[0]
Nyaw_valid = df.loc[(cond0 & cond1 & cond3),:].shape[0]
# filling in the weekly availabiliy as 1/0 based on number of points
weekly_avail = []

import more_itertools
step= 144
length = 144
idx = cond1[cond1==True].index
N_win = np.int64(len(df.loc[cond1,'WS_free_avg'])/step)
window = np.transpose(list(more_itertools.windowed(df.loc[cond1,'WS_free_avg'], n=length, fillvalue=np.nan, step=step)))
condn = np.transpose(list(more_itertools.windowed(np.array(cond0[idx] & cond1[idx] & cond2[idx] & cond4[idx]), n=length, fillvalue=np.nan, step=step))).astype(bool)
for i in np.arange(N_win):
    daily_avail = window[condn[:,i],i].shape[0]/length
    if daily_avail >= 0.6:
        weekly_avail.append(1)
    else:
        weekly_avail.append(0)

#%% Plots

# Plot all
# for i in range(len(df.columns)):
#     # date-ws_free_avg
#     fig = plt.figure(figsize =  (20, 8)) 
#     ax = fig.add_subplot(111)
#     ax.plot(df.loc[ cond1,df.columns[0]],df.loc[cond1,i],'.',color = 'gray',label = 'Invalid data');
#     ax.plot(df.loc[ cond1&cond2,df.columns[0]],df.loc[cond1&cond2,i],'.',label = 'Valid data');
#     ax.set_xlabel(df.columns[0],labelpad=40,weight= 'bold')
#     ax.set_ylabel(df.columns[i],labelpad=40,weight= 'bold')
#     date_form = DateFormatter("%d/%m")
#     ax.xaxis.set_major_formatter(date_form)
#     # ax.set_xlim([ dt.datetime.strptime(dt_start, '%Y-%m-%d %H:%M:%S'), dt.datetime.strptime(dt_end, '%Y-%m-%d %H:%M:%S')])
#     # ax.set_ylim([0,25])
#     ax.legend()
#     # plt.savefig(r'Z:\Projekte\109797-TestfeldBHV\30_Technical_execution_Confidential\TP3\AP2_Aufbau_Infrastruktur\Infrastruktur_Windmessung\02_Equipment\02_iSpin\Data\QM\TS_' + path[155:165]+'_'+df.columns[i]+'.png',bbox_inches='tight')


fig,ax = plt.subplots(4,1, figsize =  (10, 10),sharex=True) 
ax[0].plot(df.loc[ cond1&cond2,df.columns[0]],df.loc[cond1&cond2,'WS_free_avg'],'.',label = 'Valid data');
ax[0].set_xlabel('date',labelpad=40,weight= 'bold')
ax[0].set_ylabel("WS_free_avg [m/s]",labelpad=40,weight= 'bold')
date_form = DateFormatter("%d/%m")
ax[0].xaxis.set_major_formatter(date_form)
ax[0].set_xlim([ dt.datetime.strptime(dt_start, '%Y-%m-%d %H:%M:%S'), dt.datetime.strptime(dt_end, '%Y-%m-%d %H:%M:%S')])
ax[0].set_ylim([0,25])

ax[1].plot(df.loc[ cond1&cond2,df.columns[0]],df.loc[cond1&cond2,'YA_corr_avg'],'.',label = 'Valid data');
ax[1].set_xlabel('date',labelpad=40,weight= 'bold')
ax[1].set_ylabel("YA_corr_avg [$^o$]",labelpad=40,weight= 'bold')
date_form = DateFormatter("%d/%m")
ax[1].xaxis.set_major_formatter(date_form)
ax[1].set_xlim([ dt.datetime.strptime(dt_start, '%Y-%m-%d %H:%M:%S'), dt.datetime.strptime(dt_end, '%Y-%m-%d %H:%M:%S')])
ax[1].set_ylim([-45,45])

ax[2].plot(df.loc[ cond1&cond2,df.columns[0]],df.loc[cond1&cond2,'ARS_avg'],'.',label = 'Valid data');
ax[2].set_xlabel('date',labelpad=40,weight= 'bold')
ax[2].set_ylabel("ARS_avg [$^o$/s]",labelpad=40,weight= 'bold')
date_form = DateFormatter("%d/%m")
ax[2].xaxis.set_major_formatter(date_form)
ax[2].set_xlim([ dt.datetime.strptime(dt_start, '%Y-%m-%d %H:%M:%S'), dt.datetime.strptime(dt_end, '%Y-%m-%d %H:%M:%S')])
ax[2].set_ylim([0,55])

ax[3].plot(df.loc[ cond1,df.columns[0]],df.notnull().values[cond1,1]*100,'.',color = 'limegreen', label = 'Valid data');
ax[3].plot(df.loc[ cond1&(df.isnull().values[:,1]),df.columns[0]],df.notnull().values[cond1&(df.isnull().values[:,1]),1]*100,'.',color = 'red', label = 'invalid data');    
ax[3].set_xlabel('date',labelpad=40,weight= 'bold')
ax[3].set_ylabel("Availability [%]",labelpad=40,weight= 'bold')
date_form = DateFormatter("%d/%m")
ax[3].xaxis.set_major_formatter(date_form)
ax[3].set_xlim([ dt.datetime.strptime(dt_start, '%Y-%m-%d %H:%M:%S'), dt.datetime.strptime(dt_end, '%Y-%m-%d %H:%M:%S')])
ax[3].set_ylim([-5,110])
plt.xlabel('date',labelpad=10,weight= 'bold')
plt.subplots_adjust(wspace=0, hspace=0.1)
plt.savefig(r'Z:\Projekte\109797-TestfeldBHV\30_Technical_execution_Confidential\TP3\AP2_Aufbau_Infrastruktur\Infrastruktur_Windmessung\02_Equipment\02_iSpin\Data\10min_Data\QM\\'+dt_start[0:10]+'_'+dt_end[0:10]+'.png',
            bbox_inches='tight',dpi = 100)
plt.show()

## Append the quality check at the end of verification
# extracting the worksheet 
import openpyxl as opl
xl_filepath = r'z:\Projekte\109797-TestfeldBHV\30_Technical_execution_Confidential\TP3\AP2_Aufbau_Infrastruktur\Infrastruktur_Windmessung\02_Equipment\02_iSpin\Documentation\Logbook_iSpin.xlsx'
wb = opl.load_workbook(xl_filepath)
#  grab the active worksheet
ws = wb['Datenabholung']
# grab the workshet title
ws_title = ws.title
print('[{}] - Active sheet title: {} \n'.format(now(), ws_title))

# appending the worksheet
Str = '{} - {}'.format(dt_start[0:10], dt_end[0:10])
purpose = ('{}'.format(input('Please enter purpose: 0-Überwachung/Observation, 1-Datenabholung (default): ')))
if purpose == '0':
    pStr = 'Überwachung'
elif purpose == '1':
    pStr = 'Datenabholung'
else:
    pStr = 'Datenabholung'
    print('[{}] - Input not in the list! Assumed purpose= {} \n'.format(now(), pStr))

test_Data = ('Data {}'.format(input('Please enter OK or not OK plus any comments:')))
appData = [today(), pStr, Str, test_Data,Npts,Nvalid,Nws_valid, Nwt_valid, Nyaw_valid]
appData.extend(np.transpose(weekly_avail))
# target = wb.copy_worksheet(ws) % making a copy of existing worksheet

# iteration to find the last row with values in it
nrows = ws.max_row
lastrow = 0
if nrows > 1000:
    nrows = 1000
while True:
    if ws.cell(nrows, 3).value != None:
        lastrow = nrows
        break
    else:
        nrows -= 1

    # appending to the worksheet and saving it
if AppendLog==1: # if AppendLog is wished at start
    count=0
    # iterate over all entries in appData array which contains variables for appending the excel
    for ncol, entry in enumerate(appData,start=1):
        # print(ncol, entry)
        ws.cell(row=1+nrows, column=ncol, value=entry)
        count += 1
    print('[{}] - No. of entries made: {} \n'.format(now(), count))
    wb.save(xl_filepath) # file should be closed to save
    print('[{}] - Changes saved to: {} \n'.format(now(), ws_title))
else:
    print('[{}] - Changes not saved to: {} \n'.format(now(), ws_title))
    

## References:
    # https://realpython.com/openpyxl-excel-spreadsheets-python/
    
## Links to data and logbook
# z:\Projekte\109797-TestfeldBHV\30_Technical_execution_Confidential\TP3\AP2_Aufbau_Infrastruktur\Infrastruktur_Windmessung\02_Equipment\04_Nacelle-Lidars-Inflow_CONFIDENTIAL\30_Data\Logbook_NacelleLidars.xlsx'




