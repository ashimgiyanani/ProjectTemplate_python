# -*- coding: utf-8 -*-
"""
Created on Wed Mar  3 19:41:21 2021

@author: giyash
"""
# Steps to include in the script:
# import all the channels from OneDAS [##### 100%]
# arrange them into one database and one datetime format [#### 80%]
# perform a standardized check using a function definition [0%]
# perform an advanced check using a function definition [0%]
# write the results to an excel file which serves as a logbook for all the sensors [## 40%]
# save the pdf report generated on the Z drive [# 20%]
# send the report to all recipents [0%]
 
#%% user modules
import pandas as pd
import glob
import os
import numpy as np
import sys
sys.path.append(r"../../userModules")
# sys.path.append(r"../fun") # change to this when sharing the data
sys.path.append(r"../../OneDasExplorer/Python Connector")

import runpy as rp
from FnWsRange import *
import matplotlib.pyplot as plt

from csv import writer
import matlab2py as m2p
from pythonAssist import *
import tikzplotlib as tz
from AD180 import AD180
from datetime import datetime
import re

from odc_exportData import odc_exportData
from FnImportOneDas import FnImportOneDas
from datetime import *

#%% user definitions
tiny = 12
Small = 14
Medium = 16
Large = 18
Huge = 22
plt.rc('font', size=Small)          # controls default text sizes
plt.rc('axes', titlesize=Small)     # fontsize of the axes title
plt.rc('axes', labelsize=Large)    # fontsize of the x and y labels
plt.rc('xtick', labelsize=Small)    # fontsize of the tick labels
plt.rc('ytick', labelsize=Small)    # fontsize of the tick labels
plt.rc('legend', fontsize=Medium)    # legend fontsize
plt.rc('figure', titlesize=Huge)  # fontsize of the figure title


#%% import data from OneDAS
begin = datetime(2021, 6, 13, 0, 0, tzinfo=timezone.utc)
end   = datetime(2021, 6, 16, 0, 0, tzinfo=timezone.utc)
sampleRate = 600
AppendLog = 0

# must all be of the same sample rate
channel_paths = [
    '/AIRPORT/AD8_PROTOTYPE/GENERAL_DAQ/M0000_V1/600 s_mean',
    '/AIRPORT/AD8_PROTOTYPE/GENERAL_DAQ/P0420_RotorSpdRaw/600 s_mean',
    '/AIRPORT/AD8_PROTOTYPE/GENERAL_DAQ/M0010_V2/600 s_mean',
    '/AIRPORT/AD8_PROTOTYPE/GENERAL_DAQ/M0020_V3/600 s_mean',
    '/AIRPORT/AD8_PROTOTYPE/GENERAL_DAQ/M0030_V4/600 s_mean',
    '/AIRPORT/AD8_PROTOTYPE/GENERAL_DAQ/M0040_V5/600 s_mean',
    '/AIRPORT/AD8_PROTOTYPE/GENERAL_DAQ/M0050_V6/600 s_mean',
    '/AIRPORT/AD8_PROTOTYPE/GENERAL_DAQ/M0060_Precipitation/600 s_mean',
    '/AIRPORT/AD8_PROTOTYPE/GENERAL_DAQ/M0070_D1/600 s_mean_polar',
    '/AIRPORT/AD8_PROTOTYPE/GENERAL_DAQ/M0100_D4/600 s_mean_polar',
    '/AIRPORT/AD8_PROTOTYPE/GENERAL_DAQ/M0110_D5/600 s_mean_polar',
    '/AIRPORT/AD8_PROTOTYPE/GENERAL_DAQ/M0200_B1/600 s_mean',
    '/AIRPORT/AD8_PROTOTYPE/GENERAL_DAQ/M0210_B2/600 s_mean',
    '/AIRPORT/AD8_PROTOTYPE/GENERAL_DAQ/M0220_T1/600 s_mean',
    '/AIRPORT/AD8_PROTOTYPE/ISPIN/SaDataValid/600 s',
    '/AIRPORT/AD8_PROTOTYPE/ISPIN/WS_free_avg/600 s',
    '/AIRPORT/AD8_PROTOTYPE/ISPIN/DataOK/600 s',
    '/AIRPORT/AD8_PROTOTYPE/LIDAR_2020/BlackTC_110m_HWS_hub/600 s',
    '/AIRPORT/AD8_PROTOTYPE/LIDAR_2020/BluePO_110m_HWS_hub/600 s',
    '/AIRPORT/AD8_PROTOTYPE/LIDAR_2020/GreenPO_110m_HWS_hub/600 s',
    '/AIRPORT/AD8_PROTOTYPE/LIDAR_2020/BlackTC_110m_HWS_hub_availability/600 s',
    '/AIRPORT/AD8_PROTOTYPE/LIDAR_2020/BluePO_110m_HWS_hub_availability/600 s',
    '/AIRPORT/AD8_PROTOTYPE/LIDAR_2020/GreenPO_110m_HWS_hub_availability/600 s',
    '/AIRPORT/AD8_PROTOTYPE/WIND_CUBE/WC_115m_Wind_Speed/600 s'
                ]
# Provide the names that you want to give to the channel paths
ch_names = [
        'v1',
        'omega',
        'v2',
        'v3',
        'v4',
        'v5',
        'v6',
        'prec',
        'd1',
        'd4',
        'd5',
        'b1',
        'b2',
        'T1',
        's_valid',
        's_V', 
        's_ok',
        'btc_v110',
        'bpo_v110', 
        'gpo_v110', 
        'btc_Av110', 
        'bpo_Av110', 
        'gpo_Av110', 
        'wc_v115'
        ]

target_folder = r"../data"
tstart = datetime.strptime('2021-09-13_00-00-00', '%Y-%m-%d_%H-%M-%S') # Select start date in the form yyyy-mm-dd_HH-MM-SS
# funktioniert
tend = tstart + timedelta(days=7) # Select start date in the form yyyy-mm-dd_HH-MM-SS
odcData, pdData, t = FnImportOneDas(tstart, tend, channel_paths, ch_names, sampleRate, target_folder)

index = pd.date_range(tstart, periods=7, freq='D')
sensors = ['wt', 'ispin', 'btc', 'bpo', 'gpo', 'metmast', 'sonics', 'windcube']
weekly_avail = pd.DataFrame(index=index, columns=sensors)
weekly_avail = weekly_avail.fillna(0)

#%% Check the availability of the wind turbine
Npts = len(pdData.omega)
cond0 = pdData.omega.notnull()
cond1 = (pdData.omega) > 1
Avail = (cond0 & cond1).astype('uint8')
wt_avail = []
avail= np.nan*np.ones(int(Npts/144), dtype=np.float16, )
dt = np.zeros((int(Npts/144),1), dtype=object)
Nvalid = np.sum(Avail)
Av_pct = np.round(Nvalid/Npts * 100, decimals=1)

for i in np.arange(int(Npts/144)):
    avail[i] = np.round(Avail[144*i:144*(i+1)].mean(axis=0), decimals=2)
    dt[i] = t[144*i].date().strftime('%d/%m/%Y')
    if avail[i] >= 0.25:
        wt_avail.append(1)
    else:
        wt_avail.append(0)
    # del(avail)
weekly_avail['wt'] = wt_avail

#%% Check the availability of iSpin
## Filtering conditions
cond0 = pdData.s_V.notnull() # non-zero values
cond1 = (pdData.t>=tstart)&(pdData.t<=tend) # time interval filtering
cond2 = (pdData.s_valid==True) # 95% availability of 10 Hz data, wind vector +-90° in front of turbine within 10 min 
cond3 = (pdData.s_ok==True)  # data.TotalCountNoRotation=0, 95% availability, min & max rotor rpm, avg rotor rpm, free wind speed > 3.5 m/s, sample ID!= 0
cond4 = (pdData.s_V > 0) & (pdData.s_V < 50) # wind speed physical limits

## extra parameters for logbook
# no. of pts during the last week 
Npts = pdData.loc[(cond0 & cond1),'s_V'].shape[0]
Nvalid = pdData.loc[(cond0 & cond1 & cond2),'s_V'].shape[0]
Nws_valid = pdData.loc[(cond0 & cond1 & cond2),'s_V'].shape[0]
Nwt_valid = pdData.loc[(cond0 & cond1 & cond2 & cond3),'s_V'].shape[0]
Nyaw_valid = pdData.loc[(cond0 & cond1 & cond3),'s_V'].shape[0]
# filling in the weekly availabiliy as 1/0 based on number of points
ispin_avail = []

import more_itertools
step= 144
length = 144
idx = cond1[cond1==True].index
N_win = np.int64(len(pdData.loc[cond1,'s_V'])/step)
window = np.transpose(list(more_itertools.windowed(pdData.loc[cond1,'s_V'], n=length, fillvalue=np.nan, step=step)))
condn = np.transpose(list(more_itertools.windowed(np.array(cond0[idx] & cond1[idx] & cond2[idx] & cond4[idx]), n=length, fillvalue=np.nan, step=step))).astype(bool)
for i in np.arange(N_win):
    daily_avail = window[condn[:,i],i].shape[0]/length
    if daily_avail >= 0.6:
        ispin_avail.append(1)
    else:
        ispin_avail.append(0)
weekly_avail['ispin'] = ispin_avail

from matplotlib.dates import DateFormatter
from matplotlib.ticker import StrMethodFormatter
fig,ax = plt.subplots(1,1, figsize =  (10, 4),sharex=True)
ax.plot(pdData.t, pdData.s_V, 'k.', lw=0.5, label='all') 
ax.plot(pdData.t[(cond0 & cond1 & cond4)] , pdData.s_V[(cond0 & cond1 & cond4)] ,'.',label = 'Valid data')
ax.set_xlabel('date')
ax.set_ylabel("WS_free_avg [m/s]")
date_form = DateFormatter("%d/%m")
ax.xaxis.set_major_formatter(date_form)
ax.set_xlim([ datetime.strptime(str(tstart), '%Y-%m-%d %H:%M:%S'), datetime.strptime(str(tend), '%Y-%m-%d %H:%M:%S')])
ax.set_ylim([-2,25])
ax.legend()

#%% check the availability of Nacelle Lidars
dev = ['btc', 'bpo', 'gpo']
Npts, Nvalid, Nws_valid = [], [], []
for i in range(len(dev)):
    avail = []
    param = dev[i] + '_v110'
    param_Av = dev[i] + '_Av110'
    ## Filtering conditions
    cond0 = pdData[param].notnull()
    cond3 = (pdData[param_Av]>=0.0)
    # Extra parameters for logbook
    Npts.append(pdData[param].shape[0])
    Nvalid.append(pdData[param].loc[cond0].shape[0])
    Nws_valid.append(pdData[param].loc[cond0 & cond3].shape[0])
    # filling in the weekly availabiliy as 1/0 based on number of points
    import more_itertools
    step= 144
    length = 144
    N_win = np.int64(len(pdData.loc[:,param])/step)
    window = np.transpose(list(more_itertools.windowed(pdData.loc[:,param], n=length, fillvalue=np.nan, step=step)))
    condn = np.transpose(list(more_itertools.windowed(np.array(cond0 & cond2 & cond3), n=length, fillvalue=np.nan, step=step))).astype(bool)
    for j in np.arange(N_win):
        daily_avail = window[condn[:,j],j].shape[0]/length
        print('{:.1f}'.format(daily_avail))
        if daily_avail > 0.3:
            avail.append(1)
        else:
            avail.append(0)
    weekly_avail[dev[i]] = avail

#%% check the availability of the metmast sensors
param = ['v1','v2','v3','v4','v5','v6','prec','d1','d4','d5','b1','b2','T1']
# weekly_avail = pd.DataFrame(index=index, columns=param)

for i in range(len(param)):
    avail=[]
    # Filtering conditions
    cond0 = pdData[param[i]].notnull()
    # Extra parameters for logbook
    Npts.append(pdData[param[i]].shape[0])
    Nvalid.append(pdData[param[i]].loc[cond0].shape[0])
    Nws_valid.append(pdData[param[i]].loc[cond0 & cond3].shape[0])
    # filling in the weekly availabiliy as 1/0 based on number of points
    step= 144
    length = 144
    N_win = np.int64(len(pdData.loc[:,param[i]])/step)
    window = np.transpose(list(more_itertools.windowed(pdData.loc[:,param[i]], n=length, fillvalue=np.nan, step=step)))
    condn = np.transpose(list(more_itertools.windowed(np.array(cond0), n=length, fillvalue=np.nan, step=step))).astype(bool)
    for j in np.arange(N_win):
        daily_avail = window[condn[:,j],j].shape[0]/length
        print('{:.1f}'.format(daily_avail))
        if daily_avail > 0.3:
            avail.append(1)
        else:
            avail.append(0)
    weekly_avail[param[i]] = avail
    del avail
threshold = 0.25
weekly_avail['metmast'] = [int(round(x-threshold + 0.5)) for x in weekly_avail.loc[:, param].mean(axis=1)]

#%% check the availability of sonic anemometers
channel_paths =     [
    '/AIRPORT/AD8_PROTOTYPE/METMAST_EXTENSION/gill_115_u/20 Hz',
    '/AIRPORT/AD8_PROTOTYPE/METMAST_EXTENSION/gill_55_u/20 Hz',
    '/AIRPORT/AD8_PROTOTYPE/METMAST_EXTENSION/thies_25_Vx/20 Hz'
    ]
ch_names =
odcData, pdData, t = FnImportOneDas(tstart, tend, channel_paths, ch_names, sampleRate, target_folder)

avail=[]
# Filtering conditions
cond0 = pdData[param[i]].notnull()
# Extra parameters for logbook
Npts.append(pdData[param[i]].shape[0])
Nvalid.append(pdData[param[i]].loc[cond0].shape[0])
Nws_valid.append(pdData[param[i]].loc[cond0 & cond3].shape[0])
# filling in the weekly availabiliy as 1/0 based on number of points
step= 144
length = 144
N_win = np.int64(len(pdData.loc[:,param[i]])/step)
window = np.transpose(list(more_itertools.windowed(pdData.loc[:,param[i]], n=length, fillvalue=np.nan, step=step)))
condn = np.transpose(list(more_itertools.windowed(np.array(cond0), n=length, fillvalue=np.nan, step=step))).astype(bool)
for j in np.arange(N_win):
    daily_avail = window[condn[:,j],j].shape[0]/length
    print('{:.1f}'.format(daily_avail))
    if daily_avail > 0.3:
        avail.append(1)
    else:
        avail.append(0)
weekly_avail[param[i]] = avail


sys.exit('Manual stop')
#%% write the turbine availability to an excel file
import openpyxl as opl
xl_filepath = r'C:\Users\giyash\Documents\trial.xlsx'
wb = opl.load_workbook(xl_filepath)
#  grab the active worksheet
ws = wb['Sheet1']
# grab the workshet title
ws_title = ws.title
print('Active sheet title: {} \n'.format(ws_title))

# appending the worksheet
# iteration to find the last row with values in it
nrows = ws.max_row
lastrow = 0
if nrows > 1000:
    nrows = 1000
while True:
    if ws.cell(nrows, 3).value != None:
        lastrow = nrows
        break
    else:
        nrows -= 1

# appending to the worksheet and saving it
avgStr = '{} - {}'.format(tstart, tend)
test_Data = ('Data {}'.format(input('Please enter OK or not OK plus any comments:')))
appData = [today(), 'WT_Availability', 'QC check', avgStr,  test_Data, Npts, Nvalid, Av_pct ]
appData.extend(np.transpose(daily_avail))
# target = wb.copy_worksheet(ws) % making a copy of existing worksheet

if AppendLog==1: # if AppendLog is wished at start
    count=0
    # iterate over all entries in appData array which contains variables for appending the excel
    for ncol, entry in enumerate(appData,start=1):
        # print(ncol, entry)
        ws.cell(row=1+nrows, column=ncol, value=entry)
        count += 1
    print('[{}] - No. of entries made: {} \n'.format(now(), count))
    wb.save(xl_filepath) # file should be closed to save
    print('[{}] - Changes saved to: {} \n'.format(now(), ws_title))
else:
    print('[{}] - Changes not saved to: {} \n'.format(now(), ws_title))

#%% publish a report of Wind turbine availability
from reportlab.platypus import SimpleDocTemplate
from reportlab.platypus import Paragraph, Spacer, Table, Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.lib.units import inch

styles = getSampleStyleSheet()
report = SimpleDocTemplate("../results/TurbineReport.pdf")
# create a Title
report_title = Paragraph("Testfeld BHV: AD8-180 turbine Availability report", styles["h1"])

# create a Table
# generate a table borders
table_style = [('GRID', (0,0), (-1,-1), 1, colors.black)]
# add data
table_data = []
for i in range(len(dt)):
    table_data.append([dt[i], avail[i], daily_avail[i]])
report_table = Table(data=table_data, style=table_style, hAlign="LEFT")

report.build([report_title, report_table])

#%% Removed code
# path = r'c:\Users\giyash\OneDrive - Fraunhofer\Python\Data\OneDAS_2019-01-01T00-00_600_s_56c64a34'
# filename = glob.glob(path+'\AIRPORT_AD8_PROTOTYPE*.csv')
# df=pd.read_csv(filename[0], sep = ';',decimal = ',',header=9, skiprows=[10,11], usecols=[0,1,2], names=['index','cupWs', 'omega'])
# time = [begin + timedelta(seconds=i/sampleRate) for i in range(len(df['index']))]

# cond0 = pdData.omega.notnull()
# cond1 = pd.to_numeric(pdData.omega) > 0
# Avail = (cond0 & cond1).astype('uint8')
# daily_avail = []
# for i in np.arange(int(len(pdData.omega)/144)):
#     avail = Avail[144*i:144*(i+1)].mean()
#     if avail > 0.25:
#         daily_avail.append(1)
#     else:
#         daily_avail.append(0)
