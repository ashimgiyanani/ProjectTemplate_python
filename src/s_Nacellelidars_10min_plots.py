# -*- coding: utf-8 -*-
"""
Created on Wed Jul 29 07:57:54 2020

@author: papalk
"""
# purpose: perform quality checks of the Nacelle Lidar data
# logbook
# 10.02.2021: handover from Alkistis
# 10.02.2021: changed df.ix[...,i] to df.loc[..., df.columns[i]]
# 16.02.2021: added the functionality to add details to the logbook automatically, added shading=flat
# 26.05.2021: the last 2 days of the week always = 0 irrespective of the data -> corrected using the moving windows algorithm

# import os, shutil
# rawPath = r'\\ensyno.iwes.fraunhofer.de\Daten\RAW'
# destPath = r'z:\Projekte\109797-TestfeldBHV\30_Technical_execution_Confidential\TP3\AP2_Aufbau_Infrastruktur\Infrastruktur_Windmessung\02_Equipment\04_Nacelle-Lidars-Inflow_CONFIDENTIAL\30_Data\BlackTC'
# open(rawPath)

import sys
import datetime as dt
import matplotlib as matplotlib
import matplotlib.pyplot as plt
import matplotlib.pylab as pylab
from matplotlib.dates import DateFormatter
from matplotlib.ticker import StrMethodFormatter
import numpy as np
import pandas as pd
import xarray as xr
from mpl_toolkits.mplot3d import Axes3D
import glob

# local defined modules
def now():
    from datetime import datetime
    now=datetime.now().strftime("%H:%M:%S")
    return now

def today():
    from datetime import datetime
    today =  datetime.today().strftime("%d.%m.%Y")
    return today

sys.path.insert(1, r'../fun')

plt.style.use('seaborn-whitegrid')
SMALL_SIZE = 22
MEDIUM_SIZE = 22
BIGGER_SIZE = 22
AppendLog = 1 # appends into  the excel file
SpecialPlots = 1 # plots hws hub time series

plt.rc('font', size=SMALL_SIZE,weight = 'bold')          # controls default text sizes
plt.rc('axes', titlesize=SMALL_SIZE)     # fontsize of the axes title
plt.rc('axes', labelsize=BIGGER_SIZE)    # fontsize of the x and y labels
plt.rc('xtick', labelsize=MEDIUM_SIZE)    # fontsize of the tick labels
plt.rc('ytick', labelsize=MEDIUM_SIZE)    # fontsize of the tick labels
plt.rc('legend', fontsize=MEDIUM_SIZE)    # legend fontsize
plt.rc('figure', titlesize=BIGGER_SIZE)  # fontsize of the figure title
plt.rc('figure', figsize =  (8, 8))

#%% Import data - Set import parameters

# Change parameters to select the data
device = 'GreenPO' #Select between GreenPO, BluePO, BlackTC
dt_start ='2021-09-06_00-00-00' # Select start date in the form yyyy-mm-dd_HH-MM-SS
param = 'HWS hub'


dt_end = dt.datetime.strptime(dt_start, '%Y-%m-%d_%H-%M-%S')  + dt.timedelta(days=7)# Select end date in the form yyyy-mm-dd_HH-MM-SS
dt_end = dt_end.strftime('%Y-%m-%d_%H-%M-%S')
i_RaworAverage = 1 #Select if you want to look into raw (0) or average data (1)
devices = ["BlackTC","BluePO","GreenPO"]
# Set path
path = r'Z:\Projekte\109797-TestfeldBHV\30_Technical_execution_Confidential\TP3\AP2_Aufbau_Infrastruktur\Infrastruktur_Windmessung\02_Equipment\04_Nacelle-Lidars-Inflow_CONFIDENTIAL\30_Data\\'

df_all = {}
for device in devices:
    # import data
    from f_import_nacelle_lidar_data import f_import_nacelle_lidar_data
    df_all[device] = f_import_nacelle_lidar_data(path,device,i_RaworAverage,dt_start,dt_end)
    

#%%Plots
if i_RaworAverage:
    
    # Plot HWS Low
    i_range = 110
    weekly_avail = []
    Npts, Nvalid, Nws_valid = [], [], []
    fig,ax = plt.subplots(3,1, figsize =  (10, 15),sharex=True) 
    for k,n in enumerate(df_all):
        df = df_all[n]
        ## Filtering conditions
        cond0 = df[param].notnull()
        df['CNR2'] = pd.to_numeric(df['CNR2'], errors='coerce')
        df['CNR3'] = pd.to_numeric(df['CNR3'], errors='coerce')
        cond2 = (df['CNR2']>-33) & (df['CNR3']>-33)
        cond3 = (df['HWS low Availability']>=0.0)
        cond4 = (df['Distance']==i_range)
        # Extra parameters for logbook
        Npts.append(df[param].shape[0])
        Nvalid.append(df.loc[cond0 & cond2 & cond3].shape[0])
        Nws_valid.append(df[param].loc[cond0 & cond2 & cond3].shape[0])
        # filling in the weekly availabiliy as 1/0 based on number of points
        import more_itertools
        step= 144*N_dist[k]
        length = 144*N_dist[k]
        N_win = np.int64(len(df.loc[:,param])/step)
        window = np.transpose(list(more_itertools.windowed(df.loc[:,param], n=length, fillvalue=np.nan, step=step)))
        condn = np.transpose(list(more_itertools.windowed(np.array(cond0 & cond2 & cond3), n=length, fillvalue=np.nan, step=step))).astype(bool)
        for i in np.arange(N_win):
            daily_avail = window[condn[:,i],i].shape[0]/length
            print('{:.1f}'.format(daily_avail))
            if daily_avail > 0.3:
                weekly_avail.append(1)
            else:
                weekly_avail.append(0)


        if len(df)>0:
                 # 10 min
                df_CNR = df.loc[cond2,:]
                for i in range(len(df.columns)):
                  if df.columns[i] == param: # Change this if you want to plot another parameter
                    # date-ws_free_avg
                    # ax = fig.add_subplot(111)
                    ax[k].plot(df.loc[cond2&
                                     cond3&cond4,'Date and Time'],
                               df.loc[cond2&
                                     cond3&cond4,df.columns[i]],
                               '.',color = n[:-2].lower());
                    # ax.plot(df_CNR.ix[(df_CNR['HWS low Availability']<0.7),0],df_CNR.ix[(df_CNR['HWS low Availability']<0.7),i],'.',color = 'gray',label = 'Invalid data');
                    # ax[k].set_xlabel(df.columns[0],labelpad=40,weight= 'bold')
                    ax[k].set_ylabel(df.columns[i],labelpad=10,weight= 'bold')
                    date_form = DateFormatter("%d/%m")
                    ax[k].xaxis.set_major_formatter(date_form)
                    ax[k].annotate(n, xy=(0.98, 0.98), xycoords='axes fraction',
                                size=22, ha='right', va='top',
                                bbox=dict(boxstyle='round', fc='w',alpha=0.6))
                    ax[k].set_xlim([ dt.datetime.strptime(dt_start, '%Y-%m-%d_%H-%M-%S'), dt.datetime.strptime(dt_end, '%Y-%m-%d_%H-%M-%S')])
                    ax[k].set_ylim([0,25])
                    ax[k].grid(True,lw = 2)
    plt.xlabel(df.columns[0],labelpad=10,weight= 'bold')
    plt.subplots_adjust(wspace=0, hspace=0.1)
    plt.savefig(r'Z:\Projekte\109797-TestfeldBHV\30_Technical_execution_Confidential\TP3\AP2_Aufbau_Infrastruktur\Infrastruktur_Windmessung\02_Equipment\04_Nacelle-Lidars-Inflow_CONFIDENTIAL\30_Data\QM\10_WeeklyCheck\\'+dt_start+'_'+dt_end+'_r'+str(i_range)+'_TS.png',
                bbox_inches='tight',dpi = 100)
    plt.show()
    
#%% Plot availability   
    plt.rcParams["axes.axisbelow"] = False
    fig,axs = plt.subplots(3,1, figsize =  (10, 15),sharex=True) 
    for k,n in enumerate(df_all):
        df = df_all[n]
        ## Filtering conditions
        cond2 = (df['CNR2']>-33) & (df['CNR3']>-33)
        cond3 = (df['HWS low Availability']>=0.0)
        cond4 = (df['Distance']==i_range)

        if len(df)>0:
                 # 10 min
                df_CNR = df.loc[cond2,:]
                for i in range(len(df.columns)):
                  if df.columns[i] == param: # Change this if you want to plot another parameter
                      m_avail =pd.DataFrame(np.transpose(np.reshape(df.loc[:,'HWS low Availability'].values,
                                     (len(np.unique(df['Distance'])),int(len(df)/len(np.unique(df['Distance'])))))))
                      m_avail.index = df.loc[cond4,'Date and Time']
                      # remove duplicates in the index
                      m_avail = m_avail[~m_avail.index.duplicated()]
                      idx = pd.period_range(dt.datetime.strptime(dt_start, '%Y-%m-%d_%H-%M-%S'), 
                                            dt.datetime.strptime(dt_end, '%Y-%m-%d_%H-%M-%S'),freq = '600S')
                      m_avail = m_avail.reindex(idx.to_timestamp())
                      axs[k].pcolormesh(m_avail.index,
                                     np.unique(df['Distance']),
                                     np.transpose(m_avail.values),
                                     cmap = matplotlib.colors.ListedColormap(["red","yellow","limegreen"]),
                                     shading= 'auto',
                                     # vmin=np.transpose(m_avail.values).min(), vmax=np.transpose(m_avail.values).max(),
                                     zorder=1)
                      axs[k].set_ylabel('range [m]',labelpad=10,weight= 'bold')
                      date_form = DateFormatter("%d/%m")
                      axs[k].xaxis.set_major_formatter(date_form)
                      axs[k].grid(True,lw = 2)
                      axs[k].annotate(n, xy=(0.98, 0.98), xycoords='axes fraction',
                                size=22, ha='right', va='top',
                                bbox=dict(boxstyle='round', fc='w',alpha=0.6))
                      axs[k].set_xlim([ dt.datetime.strptime(dt_start, '%Y-%m-%d_%H-%M-%S'), dt.datetime.strptime(dt_end, '%Y-%m-%d_%H-%M-%S')])
    plt.xlabel(df.columns[0],labelpad=10,weight= 'bold')
    plt.subplots_adjust(wspace=0, hspace=0.1)
    # fig.colorbar(pc, ax=axs.ravel().tolist())
    plt.savefig(r'Z:\Projekte\109797-TestfeldBHV\30_Technical_execution_Confidential\TP3\AP2_Aufbau_Infrastruktur\Infrastruktur_Windmessung\02_Equipment\04_Nacelle-Lidars-Inflow_CONFIDENTIAL\30_Data\QM\10_WeeklyCheck\\'+dt_start+'_'+dt_end+'_Avail.png',
                bbox_inches='tight',dpi = 100)
    plt.show()

    if SpecialPlots == 1:
        df['HWS hub'] = df['HWS hub'].apply(pd.to_numeric, errors='coerce')
        cond_sp = (df['HWS hub']>=0) & (df['HWS hub']<=40)
        fig, ax = plt.subplots(1,1, figsize = (8,4), sharex=True)
        plt.plot(df['Date and Time'][cond_sp], df['HWS hub'][cond_sp], 'k.')
        plt.xlabel('Date and Time [10-min]')
        plt.ylabel('Wind speed [m/s]')
        plt.xticks(rotation=45)
            
    
else:
    # raw
    cond5 = (df['LOS index']==i_LOS)
    cond6 = (df['RWS Status']==1)
    i_LOS = 1
    i_range = 140
    for i in range(len(df.columns)):
        # date-ws_free_avg
        fig = plt.figure(figsize =  (20, 8)) 
        ax = fig.add_subplot(111)
        ax.plot(df.loc[cond5 & cond5,df.columns[i]],'.',label = 'Valid data');
        ax.plot(df.loc[cond6 & cond5,df.columns[i]],'.');
        ax.set_xlabel('index [-]',labelpad=40,weight= 'bold')
        ax.set_ylabel(df.columns[i]+str(i_LOS),labelpad=40,weight= 'bold')
        date_form = DateFormatter("%M:%S")
        ax.xaxis.set_major_formatter(date_form)
        # legend = ax.legend(frameon = 1)
        # frame = legend.get_frame()
        # frame.set_color('white')
        # plt.savefig(r'C:\Users\papalk\OneDrive - Fraunhofer\Nacelle lidars\Data\BlackTC\Plots\raw_TS_' + path[155:165]+'_'+df.columns[i]+'LOS'+str(LOS)+'_black.png',bbox_inches='tight')
    
    for i in range(len(df.columns)):
        # date-ws_free_avg
        fig = plt.figure(figsize =  (20, 8)) 
        ax = fig.add_subplot(111)
        # ax.plot(df.loc[cond5 & cond4,i],'.',label = 'Valid data');
        ax.plot(df.loc[cond6 & cond5 & cond4,df.columns[i]],'.',label = 'LOS 0');
        ax.plot(df.loc[cond6 & (df['LOS index']==i_LOS+1) & cond4,df.columns[i]],'.',label = 'LOS 1');
        ax.plot(df.loc[cond6 & (df['LOS index']==i_LOS+2) & cond4,df.columns[i]],'.',label = 'LOS 2');
        ax.plot(df.loc[cond6 & (df['LOS index']==i_LOS+3) & cond4,df.columns[i]],'.',label = 'LOS 3');
        ax.set_xlabel('index [-]',labelpad=40,weight= 'bold')
        ax.set_ylabel(df.columns[i],labelpad=40,weight= 'bold')
        date_form = DateFormatter("%d/%m %M:%S")
        ax.xaxis.set_major_formatter(date_form)
        legend = ax.legend(frameon = 1)
        frame = legend.get_frame()
        frame.set_color('white')

## Append the quality check at the end of verification
if AppendLog == 1:
    # extracting the worksheet 
    import openpyxl as opl
    xl_filepath = r'z:\Projekte\109797-TestfeldBHV\30_Technical_execution_Confidential\TP3\AP2_Aufbau_Infrastruktur\Infrastruktur_Windmessung\02_Equipment\04_Nacelle-Lidars-Inflow_CONFIDENTIAL\30_Data\Logbook_NacelleLidars.xlsx'
    wb = opl.load_workbook(xl_filepath)
    #  grab the active worksheet
    ws = wb['Datenabholung']
    # grab the workshet title
    ws_title = ws.title
    print('[{}] - Active sheet title: {} \n'.format(now(), ws_title))
    
    # appending the worksheet
    for i in range(len(devices)):
        avgStr = '{} - {}'.format(dt_start, dt_end)
        test_Data = ('Data {}'.format(input('Please enter OK or not OK plus any comments:')))
        appData = [today(), devices[i], 'collected',avgStr,  test_Data, Npts[i], Nvalid[i], Nws_valid[i] ]
        appData.extend(np.transpose(weekly_avail[7*i:7*(i+1)]))
        # target = wb.copy_worksheet(ws) % making a copy of existing worksheet
        
        # iteration to find the last row with values in it
        nrows = ws.max_row
        lastrow = 0
        if nrows > 1000:
            nrows = 1000
        while True:
            if ws.cell(nrows, 3).value != None:
                lastrow = nrows
                break
            else:
                nrows -= 1
        # appending to the worksheet and saving it
        count=0
        for ncol, entry in enumerate(appData,start=1):
            # print(ncol, entry)
            ws.cell(row=1+nrows, column=ncol, value=entry)
            count += 1
        print('[{}] - No. of entries made: {} \n'.format(now(), count))
    wb.save(xl_filepath) # file should be closed to save
    print('[{}] - Changes saved to: {} \n'.format(now(), ws_title))
else:
    print('[{}] - Check not appended into the logbook \n'.format(now()))

## References:
    # https://realpython.com/openpyxl-excel-spreadsheets-python/
    
## Links to data and logbook
# z:\Projekte\109797-TestfeldBHV\30_Technical_execution_Confidential\TP3\AP2_Aufbau_Infrastruktur\Infrastruktur_Windmessung\02_Equipment\04_Nacelle-Lidars-Inflow_CONFIDENTIAL\30_Data\Logbook_NacelleLidars.xlsx'


